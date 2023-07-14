from selenium import webdriver
from bs4 import BeautifulSoup

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException
import time
import re
import requests as rq
import os
from datetime import datetime, timedelta

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('avr.xlsx')

# Select the worksheet by name
worksheet = workbook['Sheet1']

# Set the starting and ending rows
starting_row = 2
ending_row = worksheet.max_row

# Loop through each row, starting at row 2, and append the cell text to a list
full_names = []
for row in range(starting_row, ending_row + 1):
    cell = worksheet.cell(row=row, column=2)
    # Use strip() method to remove any leading or trailing spaces from cell text
    if cell.value is not None:
        cell_text = cell.value.strip()
        full_names.append(cell_text)

index = 0
for full_name in full_names:
    if index > 4:
        print("Limitation exceeds")
        break
    
    split_names = full_name.split(" ")
    last_name = split_names[len(split_names) - 1]
    first_name = split_names[0]
    # Set up the webdriver
    driver = webdriver.Chrome("path/to/chromedriver.exe")
    # driver1 = webdriver.Chrome("path/to/chromedriver.exe")

    # Open the webpage and input data into the input fields
    driver.get("https://press.essexregister.com/ProdPRESS/Clerk/ClerkHome.aspx?op=basic")

    time.sleep(2)

    for i in driver.find_elements(By.CSS_SELECTOR, ".tabbernav li"):
        if i.text.find("By Name") != -1:
            i.click()

    print("li clicked")
    # time.sleep(2)

    InputField1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_txtLastNameTab1")))
    InputField1.send_keys(last_name)

    InputField2 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_txtFirstNameTab1")))
    InputField2.send_keys(first_name)

    wait1 = WebDriverWait(driver, 10)
    wait1.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_ddlDocTypeTab1"))).click()
    select1 = Select(wait1.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_ddlDocTypeTab1"))))
    select1.select_by_index(16)

    wait = WebDriverWait(driver, 10)
    wait.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_ddlShowRecTab1"))).click()
    select = Select(wait.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_ddlShowRecTab1"))))
    select.select_by_index(2)

    search_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_btnSearchTab1")))
    search_button.click()

    # check if there are no MORTGAGE results to show
    soup = BeautifulSoup(driver.page_source, "html.parser")
    view_count = soup.text.count("No results found")
    print(view_count)
    if view_count != 0:
        driver.quit()
        continue
    # Scrape the resulting page and extract the data you need
    print(driver.current_url)

    record_tds = soup.find_all("td", text = "MORTGAGE")
    
    # select the latest date
    ind = 0
    max = 0
    first_date = record_tds[0].find_next_siblings()
    first_date_text = first_date[3].text
    first_date_text = re.sub(r"[^0-9/]", "", first_date_text)
    st_date = datetime.strptime(first_date_text, "%m/%d/%Y")
    for record_td in record_tds:
        result_text = record_td.find_next_siblings()
        date_text = result_text[3].text
        date_text = re.sub(r"[^0-9/]", "", date_text)
        real_date = datetime.strptime(date_text, "%m/%d/%Y")
        
        if st_date < real_date:
            st_date = real_date
            max = ind
        
        ind += 1
    print("--------------------", max)
    max += 3
    view_num = "{:02d}".format(max)
    origin_view_id = "ctl00_ContentPlaceHolder1_dgdDeedMort_ctl03_btnView"
    main_view_id = origin_view_id[:41] + view_num + origin_view_id[43:]
    print(main_view_id)
    time.sleep(2)
    view_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, main_view_id)))
    view_button.click()

    driver.implicitly_wait(10)

    # Retrieve the reloaded page with the changed URL
    new_url = driver.current_url
    # page_source = driver.page_source

    print(new_url)
    driver.get(new_url)

    # headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"}

    # resp = rq.get(driver.current_url, headers = headers)
    # soup1 = BeautifulSoup(resp.content, "lxml")

    # trs = soup1.find_all('tr')
    # for tr in trs:
    #     print(tr.get_text().strip())

    time.sleep(3)

    driver.get ("https://press.essexregister.com/ProdPRESS/Clerk/ShowDetails.htm?789687")
    # time.sleep(3)

    iframe = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "InstViewerHeadFrame")))
    driver.switch_to.frame(iframe)
    
    # Download Mortgage document
    getMortgage_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "btnImage")))
    getMortgage_button.click()

    driver.switch_to.default_content()
    iframe1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "InstViewerBodyFrame")))
    driver.switch_to.frame(iframe1)

    save_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "Button_SaveImage")))
    save_button.click()
    time.sleep(10)

    folder_path = 'C:\\Users\\Jithu\\Downloads\\'

    # Set the current and new file names
    current_file_name = 'OPRSFile.pdf'
    new_file_name = full_name + "_" + 'MORTGAGE.pdf'

    # Use os.path.join() method to join the folder path with the old file name
    old_file_path = folder_path + current_file_name

    # Use os.path.join() method again to join the folder path with the new file name
    new_file_path = folder_path + new_file_name

    # Use os.rename() method to change the file name
    os.rename(old_file_path, new_file_path)

    driver.switch_to.default_content()
    iframe = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "InstViewerHeadFrame")))
    driver.switch_to.frame(iframe)

    get_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "btnInst")))
    get_button.click()
    # time.sleep(5)

    soup = BeautifulSoup(driver.page_source, "html.parser")
    related_doc_count = soup.text.count("View")
    print(related_doc_count)
    driver.quit()

    if related_doc_count == 0:
        driver.quit()
        continue
    origin_href = "//a[@href=\"javascript:__doPostBack('dgdRelatedInst$ctl02$ctl00','')\"]"
    for view_index in range(0, related_doc_count) :
        driver = webdriver.Chrome("path/to/chromedriver.exe")
        driver.get("https://press.essexregister.com/ProdPRESS/Clerk/ClerkHome.aspx?op=basic")

        time.sleep(2)

        for i in driver.find_elements(By.CSS_SELECTOR, ".tabbernav li"):
            if i.text.find("By Name") != -1:
                i.click()

        InputField1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_txtLastNameTab1")))
        InputField1.send_keys(last_name)

        InputField2 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_txtFirstNameTab1")))
        InputField2.send_keys(first_name)

        wait1 = WebDriverWait(driver, 10)
        wait1.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_ddlDocTypeTab1"))).click()
        select1 = Select(wait1.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_ddlDocTypeTab1"))))
        select1.select_by_index(16)

        wait = WebDriverWait(driver, 10)
        wait.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_ddlShowRecTab1"))).click()
        select = Select(wait.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_ddlShowRecTab1"))))
        select.select_by_index(2)

        search_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_btnSearchTab1")))
        search_button.click()
        print(driver.current_url)
        view_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, main_view_id)))
        view_button.click()

        driver.implicitly_wait(10)

        new_url = driver.current_url

        print(new_url)
        driver.get(new_url)
        time.sleep(3)

        driver.get ("https://press.essexregister.com/ProdPRESS/Clerk/ShowDetails.htm?789687")

        iframe = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "InstViewerHeadFrame")))
        driver.switch_to.frame(iframe)

        get_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "btnInst")))
        get_button.click()

        if view_index == 8:
            res = origin_href[:54] + "10" + origin_href[56:]
            origin_href = res
        else :
            num = int(origin_href[55])
            num += view_index
            if view_index >= 9:
                num -= 8

            changed_str = str(num)
            prefix = origin_href[:55]
            suffix = origin_href[56:]

            res = prefix + changed_str + suffix

        wait = WebDriverWait(driver, 10)
        elem = wait.until(EC.presence_of_element_located((By.XPATH, res)))
        elem.click()

        final_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "btnImage")))
        final_button.click()

        driver.switch_to.default_content()

        iframe1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "InstViewerBodyFrame")))
        driver.switch_to.frame(iframe1)

        save_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "Button_SaveImage")))
        save_button.click()
        time.sleep(3)

        driver.get("https://press.essexregister.com/ProdPRESS/Clerk/NavigateRecords.aspx?relateddoc=Y")
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        type_element = soup.find('td', text='Type')
        if type_element:
            # Find the next sibling td tag element and print its text.
            next_td_element = type_element.find_next('td')
            if next_td_element:
                print(next_td_element.text)
            else:
                print("There is no next <td> tag element after the 'Type' element.")
        else:
            print("There is no 'Type' element in the HTML content.")

        folder_path = 'C:\\Users\\Jithu\\Downloads\\'

        # Set the current and new file names
        current_file_name = 'OPRSFile.pdf'
        new_file_name = full_name + "_" + str(view_index) + "_" + next_td_element.text + '.pdf'

        # Use os.path.join() method to join the folder path with the old file name
        old_file_path = folder_path + current_file_name

        # Use os.path.join() method again to join the folder path with the new file name
        new_file_path = folder_path + new_file_name

        # Use os.rename() method to change the file name
        os.rename(old_file_path, new_file_path)

        time.sleep(5)
        driver.quit()